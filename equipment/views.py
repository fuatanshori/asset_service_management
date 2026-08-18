# equipment/views.py

import io
import json

import qrcode
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .forms import EquipmentFilterForm, EquipmentForm, EquipmentImportForm
from .models import Equipment


# ---------- Export helpers (dipakai bareng equipment_export & reports.report_export_full) ----------

EQUIPMENT_EXPORT_HEADERS = [
    "Nama", "Merk", "Tipe", "No. Seri", "Tahun Perolehan", "Status",
    "Lokasi", "Latitude", "Longitude", "Keterangan", "Dibuat Pada", "Diubah Pada",
]


def equipment_export_row(item):
    return [
        item.name, item.brand, item.model_type, item.serial_number, item.acquisition_year,
        item.get_status_display(), item.location_name,
        float(item.latitude) if item.latitude is not None else "",
        float(item.longitude) if item.longitude is not None else "",
        item.notes,
        timezone.localtime(item.created_at).strftime("%d-%m-%Y %H:%M"),
        timezone.localtime(item.updated_at).strftime("%d-%m-%Y %H:%M"),
    ]


# ---------- Import helpers (dipakai equipment_import) ----------

# Urutan kolom PERSIS sesuai file Excel asli (13 kolom, header baris 1).
# Beberapa kolom SENGAJA nggak dipakai buat import (photo, latitude,
# longitude, location_name, created_at, updated_at) — nggak bisa atau
# nggak perlu diimpor lewat Excel — tapi tetap didaftar di sini biar
# posisi kolom-kolom lain nggak geser pas dibaca.
EQUIPMENT_IMPORT_HEADERS = [
    "serial_number", "name", "brand", "model_type", "acquisition_year",
    "status", "notes", "photo", "latitude", "longitude", "location_name",
    "created_at", "updated_at",
]

# Nilai status di Excel dicocokkan case-insensitive ke status internal
# aplikasi. Nilai yang nggak dikenali (atau kosong) default ke Aktif.
STATUS_IMPORT_MAP = {
    "active": Equipment.STATUS_ACTIVE,
    "aktif": Equipment.STATUS_ACTIVE,
    "damaged": Equipment.STATUS_DAMAGED,
    "rusak": Equipment.STATUS_DAMAGED,
}


def _parse_import_row(row_num, row):
    """Validasi & bersihin satu baris Excel jadi dict siap dipakai
    Equipment.objects.create(). Return (data_dict, None) kalau valid,
    (None, pesan_error) kalau ada yang salah — baris itu dilewati, baris
    lain tetap diproses.

    Kolom yang diawali underscore (_photo, _latitude, dst) SENGAJA
    diterima tapi diabaikan — nggak bisa/nggak perlu diimpor lewat
    Excel (foto butuh file asli, created_at/updated_at otomatis dari
    model, dst) — cuma dijaga posisinya biar kolom lain nggak geser.

    name wajib. brand, serial_number, & acquisition_year BOLEH kosong
    (data lama sering nggak lengkap) — serial_number kosong disimpan
    sebagai None (bukan string kosong), biar banyak barang boleh
    sama-sama nggak punya nomor seri tanpa nabrak constraint unique di
    model (lihat catatan di Equipment.serial_number).

    status dibaca dari Excel (active/aktif -> Aktif, damaged/rusak ->
    Rusak) — nilai lain atau kosong default ke Aktif. Ini AMAN dipasang
    langsung meski barang belum punya jadwal maintenance: aturan "tanpa
    jadwal = Aktif" di sync_equipment_status() cuma jalan pas ada
    log/jadwal disave, bukan dicek ulang terus-menerus — jadi status
    hasil impor kesimpen apa adanya sampai memang ada aktivitas
    maintenance beneran buat barang itu."""
    values = (list(row) + [None] * len(EQUIPMENT_IMPORT_HEADERS))[: len(EQUIPMENT_IMPORT_HEADERS)]
    (
        serial_number, name, brand, model_type, acquisition_year,
        status_raw, notes, _photo, _latitude, _longitude, _location_name,
        _created_at, _updated_at,
    ) = values

    name = str(name).strip() if name not in (None, "") else ""
    brand = str(brand).strip() if brand not in (None, "") else ""
    serial_number = str(serial_number).strip() if serial_number not in (None, "") else ""

    if not name:
        return None, f"Baris {row_num}: Nama Barang kosong — dilewati."

    if acquisition_year in (None, ""):
        acquisition_year = None
    else:
        try:
            acquisition_year = int(acquisition_year)
        except (TypeError, ValueError):
            # Nangkep kasus kayak "2020.0" (angka desimal dalam bentuk
            # teks, sering dari hasil export/formula sistem lain) —
            # int() langsung nolak string berdesimal, jadi dicoba lewat
            # float() dulu sebelum beneran nyerah.
            try:
                acquisition_year = int(float(acquisition_year))
            except (TypeError, ValueError):
                acquisition_year = None  # beneran bukan angka — dikosongin, bukan nolak baris

    if serial_number:
        if Equipment.objects.filter(serial_number=serial_number).exists():
            return None, f"Baris {row_num}: Nomor Seri '{serial_number}' sudah terdaftar — dilewati."
    else:
        serial_number = None  # None (bukan ""), biar aman di kolom unique=True

    status_key = str(status_raw).strip().lower() if status_raw not in (None, "") else ""
    status = STATUS_IMPORT_MAP.get(status_key, Equipment.STATUS_ACTIVE)

    return {
        "name": name,
        "brand": brand,
        "model_type": str(model_type).strip() if model_type not in (None, "") else "",
        "serial_number": serial_number,
        "acquisition_year": acquisition_year,
        "status": status,
        "notes": str(notes).strip() if notes not in (None, "") else "",
    }, None


# ---------- Equipment ----------

@login_required
def equipment_list(request):
    form = EquipmentFilterForm(request.GET or None)
    qs = Equipment.objects.all()
    if form.is_valid():
        data = form.cleaned_data
        if data["q"]:
            q = data["q"]
            qs = qs.filter(
                Q(name__icontains=q) | Q(brand__icontains=q)
                | Q(model_type__icontains=q) | Q(serial_number__icontains=q)
            )
        if data["brand"]:
            qs = qs.filter(brand__iexact=data["brand"])
        if data["model_type"]:
            qs = qs.filter(model_type__iexact=data["model_type"])
        if data["acquisition_year"]:
            qs = qs.filter(acquisition_year=data["acquisition_year"])
        if data["status"]:
            qs = qs.filter(status=data["status"])

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    querydict = request.GET.copy()
    querydict.pop("page", None)

    context = {
        "active_page": "equipment",
        "form": form,
        "page_obj": page_obj,
        "equipment_list": page_obj.object_list,
        "filter_querystring": querydict.urlencode(),
        "brand_choices": Equipment.objects.values_list("brand", flat=True).distinct(),
        "model_type_choices": Equipment.objects.values_list("model_type", flat=True).distinct(),
        "year_choices": Equipment.objects.exclude(acquisition_year__isnull=True)
        .values_list("acquisition_year", flat=True)
        .distinct()
        .order_by("-acquisition_year"),
    }
    return render(request, "equipment/equipment_list.html", context)


def equipment_detail(request, pk):
    """Sengaja PUBLIK (tidak login_required) — halaman ini yang dituju QR
    code fisik di barang, jadi siapapun yang scan bisa langsung lihat info
    & riwayat tanpa perlu login dulu. Read-only: tombol aksi (Edit,
    Tambah, dst) di template disembunyikan untuk pengunjung yang belum
    login lewat pengecekan {% if user.is_authenticated %}."""
    item = get_object_or_404(Equipment, pk=pk)
    context = {
        "active_page": "equipment",
        "equipment": item,
        "schedule_list": item.maintenance_schedules.select_related("log").all()[:10],
    }
    return render(request, "equipment/equipment_detail.html", context)


@login_required
def equipment_add(request):
    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save()
            messages.success(request, f'Barang "{item.name}" berhasil ditambahkan.')
            return redirect("equipment_list")
    else:
        form = EquipmentForm()
    return render(
        request, "equipment/equipment_form.html", {"form": form, "active_page": "equipment", "mode": "add"}
    )


@login_required
def equipment_edit(request, pk):
    item = get_object_or_404(Equipment, pk=pk)
    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Barang "{item.name}" berhasil diperbarui.')
            return redirect("equipment_list")
    else:
        form = EquipmentForm(instance=item)
    return render(
        request,
        "equipment/equipment_form.html",
        {"form": form, "active_page": "equipment", "mode": "edit", "equipment": item},
    )


@login_required
def equipment_delete(request, pk):
    item = get_object_or_404(Equipment, pk=pk)
    if request.method == "POST":
        name = item.name
        item.delete()
        messages.success(request, f'Barang "{name}" berhasil dihapus.')
        return redirect("equipment_list")
    return render(
        request, "equipment/equipment_confirm_delete.html", {"equipment": item, "active_page": "equipment"}
    )


@login_required
def equipment_import(request):
    """Import massal Equipment dari file Excel (.xlsx) — dipakai buat
    migrasi data awal dari spreadsheet inventaris lama. Baris pertama
    dianggap header dan dilewati; urutan kolom lihat
    EQUIPMENT_IMPORT_HEADERS.

    Semua barang yang berhasil diimpor otomatis Status = Aktif (default
    field status di model, sengaja nggak di-override di sini) dan
    Dibuat/Diubah Pada = waktu import (otomatis dari auto_now_add /
    auto_now di model, nggak perlu — dan nggak bisa — diisi manual).

    Baris yang error (field wajib kosong, tahun nggak valid, nomor seri
    udah kepakai) dilewati satu-satu, bukan bikin seluruh import gagal —
    daftar barang yang berhasil & baris yang dilewati sama-sama
    ditampilin di akhir."""
    results = None
    if request.method == "POST":
        form = EquipmentImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = load_workbook(form.cleaned_data["file"], data_only=True)
                ws = wb.active
            except Exception:
                messages.error(
                    request,
                    "Gagal membaca file Excel — pastikan formatnya benar (.xlsx) dan tidak corrupt.",
                )
                return render(
                    request,
                    "equipment/equipment_import.html",
                    {"form": form, "active_page": "equipment", "results": None},
                )

            created, errors = [], []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row is None or all(cell in (None, "") for cell in row):
                    continue  # baris kosong, lewati diam-diam
                data, error = _parse_import_row(row_num, row)
                if error:
                    errors.append(error)
                    continue
                created.append(Equipment.objects.create(**data))

            results = {"created": created, "errors": errors}
            if created:
                messages.success(request, f"{len(created)} barang berhasil diimpor.")
            if errors:
                messages.warning(request, f"{len(errors)} baris dilewati — lihat detail di bawah.")
            form = EquipmentImportForm()  # form kosong lagi biar siap import berikutnya
    else:
        form = EquipmentImportForm()
    return render(
        request,
        "equipment/equipment_import.html",
        {"form": form, "active_page": "equipment", "results": results},
    )


@login_required
def equipment_search(request):
    """AJAX endpoint untuk search-as-you-type di form Jadwal (app maintenance).
    Barang yang sedang "Dalam Perbaikan" atau "Dijadwalkan" sengaja tidak
    ikut muncul, karena tidak boleh ditambahkan jadwal baru (cegah
    double-booking). Login-protected karena cuma dipakai dari dalam form
    Jadwal yang juga wajib login."""
    q = request.GET.get("q", "").strip()
    qs = Equipment.objects.exclude(status__in=[Equipment.STATUS_UNDER_REPAIR, Equipment.STATUS_SCHEDULED])
    if q:
        search_filter = (
            Q(name__icontains=q)
            | Q(brand__icontains=q)
            | Q(model_type__icontains=q)
            | Q(serial_number__icontains=q)
            | Q(location_name__icontains=q)
            | Q(notes__icontains=q)
        )
        # acquisition_year field-nya integer — icontains langsung ke
        # field angka nggak reliable di semua database (bisa error kalau
        # q bukan angka). Cuma diikutkan kalau q-nya emang murni angka.
        if q.isdigit():
            search_filter |= Q(acquisition_year=int(q))
        qs = qs.filter(search_filter)
    # "label" dipertahankan buat ngisi kotak input pas dipilih (format
    # ringkas, dipakai juga di selected_equipment_label view lain).
    # Field lainnya buat nampilin detail lebih informatif di dropdown
    # hasil pencarian — biar gampang bedain barang yang namanya mirip.
    results = [
        {
            "id": item.pk,
            "label": str(item),
            "name": item.name,
            "serial_number": item.serial_number,
            "brand": item.brand,
            "model_type": item.model_type,
            "location_name": item.location_name,
        }
        for item in qs[:10]
    ]
    return JsonResponse({"results": results})


@login_required
def equipment_export(request):
    qs = Equipment.objects.all()
    status = request.GET.get("status")
    if status:
        qs = qs.filter(status=status)

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="data_barang.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Barang"
    ws.append(EQUIPMENT_EXPORT_HEADERS)
    for item in qs:
        ws.append(equipment_export_row(item))
    wb.save(response)
    return response


@login_required
def equipment_map(request):
    items = Equipment.objects.filter(latitude__isnull=False, longitude__isnull=False)
    markers = [
        {
            "id": item.pk,
            "name": item.name,
            "lat": float(item.latitude),
            "lng": float(item.longitude),
            "status": item.get_status_display(),
            "status_color": item.status_color,
            "url": item.get_absolute_url(),
        }
        for item in items
    ]
    return render(
        request, "equipment/equipment_map.html",
        {"markers_json": json.dumps(markers), "active_page": "equipment"},
    )


def equipment_qrcode(request, pk):
    """Sengaja PUBLIK — ini gambar QR code itu sendiri, di-embed lewat
    <img> di halaman Detail Barang yang juga publik. Kalau ini
    login_required, gambarnya nggak bakal muncul buat pengunjung yang
    belum login."""
    item = get_object_or_404(Equipment, pk=pk)
    detail_url = request.build_absolute_uri(item.get_absolute_url())
    try:
        box_size = int(request.GET.get("size", 8))
    except ValueError:
        box_size = 8
    box_size = max(2, min(box_size, 20))

    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=box_size, border=2)
    qr.add_data(detail_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#14213D", back_color="white")

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return HttpResponse(buffer.getvalue(), content_type="image/png")


@login_required
def equipment_qrcode_print(request, pk):
    item = get_object_or_404(Equipment, pk=pk)
    return render(request, "equipment/equipment_qrcode_print.html", {"equipment": item})


@login_required
def equipment_qrcode_print_bulk(request):
    ids = request.GET.getlist("ids")
    items = Equipment.objects.filter(pk__in=ids)
    return render(request, "equipment/equipment_qrcode_print_bulk.html", {"equipment_list": items})


# ---------- Scan ----------

def scan_view(request):
    """Sengaja PUBLIK — halaman scan kamera dipakai di lapangan sebelum
    tentu sempat login. Link "Cari Manual" di halaman ini mengarah ke
    equipment_list yang wajib login, jadi pengunjung anonim yang pakai
    fitur itu bakal diarahkan ke halaman login dulu — itu sesuai desain."""
    return render(request, "equipment/scan.html", {"active_page": "scan"})