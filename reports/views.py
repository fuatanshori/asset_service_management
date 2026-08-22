# reports/views.py
import json
import os
from collections import OrderedDict

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.db.models import Avg, Count, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from xhtml2pdf import pisa

from equipment.analytics import get_equipment_age_analysis, get_equipment_status_breakdown
from equipment.models import Equipment
from equipment.views import EQUIPMENT_EXPORT_HEADERS, equipment_export_row
from maintenance.analytics import (
    get_cost_by_equipment,
    get_cost_by_month_multi_year,
    get_cost_summary,
    get_cost_trend,
    get_month_over_month_cost,
    get_stale_equipment,
)
from maintenance.models import MaintenanceLog, MaintenanceSchedule
from maintenance.views import SCHEDULE_EXPORT_HEADERS, schedule_export_row
from .forms import ReportFilterForm


@login_required
def report_view(request):
    filter_form = ReportFilterForm(request.GET)
    filter_form.is_valid()
    date_from = filter_form.cleaned_data.get("date_from")
    date_to = filter_form.cleaned_data.get("date_to")

    cost_trend = get_cost_trend(date_from=date_from, date_to=date_to)
    cost_by_month_multi_year = get_cost_by_month_multi_year()
    available_cost_years = sorted(cost_by_month_multi_year.keys(), reverse=True)
    age_analysis = get_equipment_age_analysis()
    stale_data = get_stale_equipment(months_threshold=6, limit=10)

    context = {
        "active_page": "report",
        "filter_form": filter_form,
        "is_filtered": bool(date_from or date_to),
        "generated_at": timezone.localtime(),
        "total_equipment": Equipment.objects.count(),
        "status_breakdown": get_equipment_status_breakdown(),
        "age_analysis": age_analysis,
        "age_buckets_json": json.dumps(age_analysis["buckets"]),
        "cost_summary": get_cost_summary(),
        "cost_summary_count": MaintenanceLog.objects.exclude(completed_date__isnull=True).count(),
        "cost_mom": get_month_over_month_cost(),
        "cost_trend": list(zip(cost_trend["labels"], cost_trend["values"])),
        # Data mentah (bukan hasil zip) buat dikonsumsi Chart.js —
        # nggak bisa pakai "cost_trend" yang udah di-zip di atas, soalnya
        # Chart.js butuh array label & array angka yang terpisah.
        "cost_trend_labels_json": json.dumps(cost_trend["labels"]),
        "cost_trend_values_json": json.dumps(cost_trend["values"]),
        "cost_by_equipment": get_cost_by_equipment(date_from=date_from, date_to=date_to, limit=10),
        "stale_equipment": stale_data["items"],
        "stale_equipment_total": stale_data["total_count"],
        # Fitur "Cek Biaya per Tahun & Bulan" — SEMUA tahun yang punya
        # data dikirim sekaligus di sini, biar ganti tahun/bulan di
        # halaman Laporan murni JS, nggak reload sama sekali.
        "available_cost_years": available_cost_years,
        "default_cost_year": available_cost_years[0] if available_cost_years else timezone.localtime().year,
        "cost_by_month_multi_year_json": json.dumps(cost_by_month_multi_year),
    }
    return render(request, "reports/report.html", context)


@login_required
def report_cost_analysis_ajax(request):
    """Endpoint JSON buat filter tanggal di section "Analisis Biaya" —
    dipanggil lewat fetch() dari JS, BUKAN navigasi/reload halaman
    biasa. Ngembaliin ringkasan (total/jumlah servis/rata-rata), chart
    Tren Biaya, DAN tabel Barang Paling Boros — semuanya dihitung dari
    queryset dasar yang sama, biar konsisten satu sama lain.

    SENGAJA cuma dengerin date_from/date_to (filter form biasa) — fitur
    "Cek Biaya per Tahun & Bulan" itu widget yang berdiri sendiri,
    nggak nyentuh endpoint ini sama sekali."""
    filter_form = ReportFilterForm(request.GET)
    filter_form.is_valid()
    date_from = filter_form.cleaned_data.get("date_from")
    date_to = filter_form.cleaned_data.get("date_to")

    logs_qs = MaintenanceLog.objects.exclude(completed_date__isnull=True)
    if date_from:
        logs_qs = logs_qs.filter(completed_date__gte=date_from)
    if date_to:
        logs_qs = logs_qs.filter(completed_date__lte=date_to)

    summary_agg = logs_qs.aggregate(total=Sum("cost"), avg=Avg("cost"), count=Count("id"))

    cost_trend_rows = list(
        logs_qs.annotate(month=TruncMonth("completed_date"))
        .values("month")
        .annotate(total_cost=Sum("cost"))
        .order_by("month")
    )

    cost_by_equipment_rows = (
        logs_qs.values("schedule__equipment__id", "schedule__equipment__name", "schedule__equipment__serial_number")
        .annotate(total_cost=Sum("cost"), service_count=Count("id"))
        .order_by("-total_cost")[:10]
    )

    return JsonResponse({
        "summary": {
            "total": float(summary_agg["total"] or 0),
            "avg_per_service": float(summary_agg["avg"] or 0),
            "service_count": summary_agg["count"] or 0,
        },
        "cost_trend_labels": [r["month"].strftime("%b %Y") for r in cost_trend_rows],
        "cost_trend_values": [float(r["total_cost"] or 0) for r in cost_trend_rows],
        "cost_by_equipment": [
            {
                "name": row["schedule__equipment__name"],
                "serial_number": row["schedule__equipment__serial_number"] or "",
                "service_count": row["service_count"],
                "total_cost": float(row["total_cost"] or 0),
            }
            for row in cost_by_equipment_rows
        ],
    })


@login_required
def report_detail_redirect(request):
    """URL lama /laporan/detail/ sekarang digabung ke /laporan/ —
    redirect biar link/bookmark lama gak 404."""
    return redirect("report")


@login_required
def report_export_full(request):
    """Excel satu file isinya banyak sheet — pakai helper baris yang sama
    persis dengan equipment_export & schedule_export (app equipment &
    maintenance), biar datanya selalu identik, gak pernah drift beda.

    Sheet "Ringkasan Barang (Digabung)" juga pakai helper grouping yang
    sama persis dengan report_export_equipment_summary (lihat
    _group_equipment_for_summary di bawah) — biar kontennya identik
    dengan versi standalone-nya, cuma beda tempat aksesnya doang."""
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="laporan_lengkap.xlsx"'

    wb = Workbook()

    # --- Sheet 1: Ringkasan ---
    ws = wb.active
    ws.title = "Ringkasan"
    status_breakdown = get_equipment_status_breakdown()
    cost_summary = get_cost_summary()
    ws.append(["Laporan Lengkap Asset Service Management System"])
    ws.append(["Digenerate pada", timezone.localtime().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    ws.append(["Total Barang", Equipment.objects.count()])
    ws.append(["Barang Aktif", status_breakdown["active"]])
    ws.append(["Barang Dijadwalkan", status_breakdown["scheduled"]])
    ws.append(["Barang Dalam Perbaikan", status_breakdown["under_repair"]])
    ws.append(["Barang Rusak", status_breakdown["damaged"]])
    ws.append([])
    ws.append(["Total Biaya Bulan Ini", cost_summary["this_month"]])
    ws.append(["Total Biaya Tahun Ini", cost_summary["this_year"]])
    ws.append(["Total Biaya Keseluruhan", cost_summary["all_time"]])
    ws.append(["Rata-rata Biaya per Servis", cost_summary["avg_per_service"]])

    # --- Sheet 2: Data Barang ---
    ws2 = wb.create_sheet("Data Barang")
    ws2.append(EQUIPMENT_EXPORT_HEADERS)
    for item in Equipment.objects.all():
        ws2.append(equipment_export_row(item))

    # --- Sheet 3: Ringkasan Barang (Digabung) ---
    # Sama persis logikanya sama report_export_equipment_summary —
    # barang dengan nama+merk+tipe+no.seri yang identik (case-insensitive)
    # digabung jadi 1 baris, kolom "Total" nunjukkin berapa unit fisik
    # yang tergabung.
    ws3 = wb.create_sheet("Ringkasan Barang (Digabung)")
    ws3.append(EQUIPMENT_EXPORT_HEADERS + ["Total"])
    groups = _group_equipment_for_summary(Equipment.objects.all())
    for data in groups.values():
        row = equipment_export_row(data["representative"])
        row.append(data["count"])
        ws3.append(row)

    # --- Sheet 4: Jadwal & Riwayat ---
    ws4 = wb.create_sheet("Jadwal & Riwayat")
    ws4.append(SCHEDULE_EXPORT_HEADERS)
    for s in MaintenanceSchedule.objects.select_related("equipment", "log").all():
        ws4.append(schedule_export_row(s))

    # --- Sheet 5: Biaya per Bulan ---
    # all_time=True (BUKAN limit_months kayak di halaman/PDF ringkas) —
    # "Semua Sekaligus" itu arsip lengkap, jadi sheet ini juga harus
    # nyakup SELURUH riwayat, biar konsisten sama sheet lain di file
    # ini (Data Barang, Jadwal & Riwayat, Barang Paling Boros — yang
    # semuanya emang udah all-time dari awal).
    ws5 = wb.create_sheet("Biaya per Bulan")
    cost_trend = get_cost_trend(all_time=True)
    if cost_trend["labels"]:
        ws5.append([f"Data dari {cost_trend['labels'][0]} sampai {cost_trend['labels'][-1]}"])
        ws5.append([])
    ws5.append(["Bulan", "Total Biaya"])
    for label, value in zip(cost_trend["labels"], cost_trend["values"]):
        ws5.append([label, value])

    # --- Sheet 6: Barang Paling Boros Biaya ---
    ws6 = wb.create_sheet("Barang Paling Boros")
    ws6.append(["Nama", "No. Seri", "Jumlah Servis", "Total Biaya"])
    for row in get_cost_by_equipment(limit=20):
        ws6.append([
            row["schedule__equipment__name"],
            row["schedule__equipment__serial_number"],
            row["service_count"],
            float(row["total_cost"] or 0),
        ])

    wb.save(response)
    return response

def _pdf_static_link_callback(uri, rel):
    """xhtml2pdf butuh PATH FILE ASLI buat baca gambar (logo, dst)
    langsung dari disk — beda sama browser biasa yang bisa akses lewat
    HTTP. {% static %} di template cuma ngehasilin URL (/static/...),
    jadi fungsi ini yang nerjemahin URL itu jadi path file beneran.
    Pola ini standar dari dokumentasi xhtml2pdf buat integrasi sama
    Django, dipasang lewat parameter link_callback= di pisa.CreatePDF().

    Coba lewat finders dulu (jalan pas DEBUG=True / belum
    collectstatic) — kalau nggak ketemu, fallback ke STATIC_ROOT
    (jalan di production, setelah collectstatic beneran ngumpulin
    semua static file ke satu folder)."""
    result = finders.find(uri.replace(settings.STATIC_URL, ""))
    if result:
        path = result[0] if isinstance(result, (list, tuple)) else result
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
    else:
        return uri  # URL absolut (http://...) — biarin apa adanya

    if not os.path.isfile(path):
        raise Exception(f"File statis nggak ketemu buat PDF: {path}")
    return path


@login_required
def report_export_pdf(request):
    """Export Laporan versi PDF — cocok buat presentasi formal (misal
    ke manajemen), beda tujuan dari export Excel yang lebih ke arah
    data mentah/kerja lanjutan.

    Pakai xhtml2pdf (render HTML -> PDF) — SENGAJA bukan WeasyPrint,
    biar nggak butuh library sistem (GTK/Pango/Cairo) yang ribet
    dipasang khususnya di Windows. xhtml2pdf murni Python, tinggal pip
    install, jalan sama persis di Windows maupun di container Linux
    tanpa perlu apt-get tambahan apapun.

    SENGAJA pakai template terpisah (report_pdf.html), BUKAN
    report.html yang biasa — baik WeasyPrint maupun xhtml2pdf sama-sama
    nggak jalanin JavaScript, jadi semua chart Chart.js (canvas) bakal
    kosong kalau dipaksa pakai template yang sama. Template PDF ini
    nunjukin data yang sama tapi dalam bentuk tabel statis, dan nggak
    nyertain widget interaktif "Cek Biaya per Tahun & Bulan" — itu
    emang dirancang buat dipakai on-screen, nggak masuk akal buat
    dokumen statis.

    Dukungan CSS xhtml2pdf lebih terbatas dari WeasyPrint (nggak ada
    CSS Paged Media standar) — makanya report_pdf.html pakai sintaks
    @frame + <pdf:pagenumber/> ala xhtml2pdf buat nomor halaman, sama
    tabel HTML asli buat kotak statistik (bukan CSS display:table)."""
    status_breakdown = get_equipment_status_breakdown()
    cost_summary = get_cost_summary()
    cost_mom = get_month_over_month_cost()
    cost_trend = get_cost_trend(limit_months=12)
    age_analysis = get_equipment_age_analysis()
    stale_data = get_stale_equipment(months_threshold=6, limit=10)

    context = {
        "generated_at": timezone.localtime(),
        "total_equipment": Equipment.objects.count(),
        "status_breakdown": status_breakdown,
        "cost_summary": cost_summary,
        "cost_mom": cost_mom,
        "cost_trend": list(zip(cost_trend["labels"], cost_trend["values"])),
        "cost_by_equipment": get_cost_by_equipment(limit=10),
        "stale_equipment": stale_data["items"],
        "stale_equipment_total": stale_data["total_count"],
        "age_analysis": age_analysis,
    }

    html_string = render_to_string("reports/report_pdf.html", context)

    response = HttpResponse(content_type="application/pdf")
    filename = f"laporan_{timezone.localdate().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html_string, dest=response, link_callback=_pdf_static_link_callback)
    if pisa_status.err:
        return HttpResponse(
            "Gagal membuat PDF — cek log server buat detail errornya.", status=500
        )
    return response


def _normalize_for_grouping(value):
    """Bikin nilai jadi case-insensitive & rapi buat keperluan
    perbandingan grouping doang — "Ventilator", "ventilator", dan
    " VENTILATOR " semuanya dianggap sama. None dan string kosong/spasi
    juga disamain jadi None, biar barang yang nggak punya nilai (misal
    nomor seri kosong) tetap konsisten dianggap "sama" sesama yang juga
    kosong. Nilai ASLI (bukan hasil normalize ini) yang tetap dipakai
    buat ditampilkan di baris Excel — normalisasi ini cuma buat nentuin
    mana yang dianggap "duplikat", bukan buat ngubah data yang tersimpan
    atau yang ditampilkan."""
    if value is None:
        return None
    value = str(value).strip().casefold()
    return value or None


def _group_equipment_for_summary(qs):
    """Kelompokkan Equipment berdasarkan (nama, merk, tipe, nomor seri)
    — barang yang persis sama di 4 field ini (case-insensitive, lihat
    _normalize_for_grouping) dianggap "duplikat" dan digabung jadi 1
    baris. Berguna khususnya buat barang generik/nggak punya nomor seri
    unik (banyak kejadian di data hasil import Excel lama), yang tanpa
    ini bakal keliatan sebagai baris berulang identik — termasuk kalau
    cuma beda huruf besar/kecil doang pas ngetik data.

    Field selain 4 itu (status, lokasi, tahun, dst) diambil dari
    ANGGOTA PERTAMA tiap kelompok (diurutkan nama lalu id) — kalau
    field itu beda-beda antar anggota kelompok yang digabung, cuma satu
    nilai representatif yang ditampilkan, bukan semuanya."""
    groups = OrderedDict()
    for item in qs.order_by("name", "pk"):
        key = (
            _normalize_for_grouping(item.name),
            _normalize_for_grouping(item.brand),
            _normalize_for_grouping(item.model_type),
            _normalize_for_grouping(item.serial_number),
        )
        if key not in groups:
            groups[key] = {"representative": item, "count": 0}
        groups[key]["count"] += 1
    return groups


@login_required
def report_export_equipment_summary(request):
    """Export ringkasan Data Barang — barang dengan nama, merk, tipe, &
    nomor seri yang sama persis digabung jadi 1 baris, dengan kolom
    "Total" (paling kanan) menunjukkan berapa banyak barang fisik yang
    digabung ke baris itu. Export terpisah dari equipment_export biasa
    (yang tetap satu baris per barang) — biar kebutuhan detail per-item
    (dipakai juga di report_export_full) tetap nggak kesentuh."""
    qs = Equipment.objects.all()
    groups = _group_equipment_for_summary(qs)

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="ringkasan_barang.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan Barang"
    ws.append(EQUIPMENT_EXPORT_HEADERS + ["Total"])
    for data in groups.values():
        row = equipment_export_row(data["representative"])
        row.append(data["count"])
        ws.append(row)
    wb.save(response)
    return response